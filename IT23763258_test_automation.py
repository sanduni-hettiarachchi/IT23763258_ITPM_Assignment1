"""
Singlish -> Sinhala Transliteration Test Automation
Target: https://www.pixelssuite.com/chat-translator
Tool: Playwright (Python)
"""

import argparse
import time
import sys

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


def parse_args():
    p = argparse.ArgumentParser(description="Singlish transliteration tester")
    p.add_argument("--excel",         required=True)
    p.add_argument("--url",           required=True)
    p.add_argument("--wait-ms",       type=int, default=8000)
    p.add_argument("--type-delay-ms", type=int, default=80)
    p.add_argument("--slow-mo-ms",    type=int, default=200)
    p.add_argument("--save-every",    type=int, default=1)
    p.add_argument("--keep-open",     action="store_true")
    p.add_argument("--headless",      action="store_true")
    return p.parse_args()


def load_workbook_data(path):
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    header_row = None
    col_map = {}

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        temp = {}
        for j, cell in enumerate(row, start=1):
            if cell is None:
                continue
            v = str(cell).strip().lower()
            if v == "input":
                temp["input"] = j
            elif "expected" in v and "output" in v:
                temp["expected"] = j
            elif "actual" in v and "output" in v:
                temp["actual"] = j
            elif v == "status":
                temp["status"] = j
        if "input" in temp and "expected" in temp:
            header_row = i
            col_map = temp
            break

    if header_row is None:
        print("WARNING: Auto-detect failed. Using hardcoded columns C,D,E,F.")
        header_row = 1
        col_map = {"input": 3, "expected": 4, "actual": 5, "status": 6}

    if "actual" not in col_map:
        col_map["actual"] = 5
        ws.cell(row=header_row, column=5, value="Actual Output")
    if "status" not in col_map:
        col_map["status"] = 6
        ws.cell(row=header_row, column=6, value="Status")

    print(f"\nHeader row   : {header_row}")
    print(f"Input  col   : {col_map['input']}  (Col {chr(64+col_map['input'])})")
    print(f"Expected col : {col_map['expected']}  (Col {chr(64+col_map['expected'])})")
    print(f"Actual col   : {col_map['actual']}  (Col {chr(64+col_map['actual'])})")
    print(f"Status col   : {col_map['status']}  (Col {chr(64+col_map['status'])})\n")

    return wb, ws, header_row, col_map


def get_data_rows(ws, header_row, col_map):
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        singlish = ws.cell(row=r, column=col_map["input"]).value
        expected = ws.cell(row=r, column=col_map["expected"]).value
        if singlish and str(singlish).strip():
            rows.append((r, str(singlish).strip(), str(expected).strip() if expected else ""))
    return rows


def setup_page(playwright, url, slow_mo, headless):
    browser = playwright.chromium.launch(
        headless=headless,
        slow_mo=slow_mo,
        args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
    )
    context = browser.new_context(
        user_agent=(
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/124.0.0.0 Safari/537.36"
        ),
        viewport={"width": 1280, "height": 800},
    )
    page = context.new_page()
    print(f"\nOpening browser: {url}\nPlease wait...\n")
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=60_000)
        time.sleep(5)
    except PlaywrightTimeout:
        print("WARNING: Page load timed out. Continuing anyway...")
        time.sleep(3)
    except Exception as e:
        print(f"ERROR loading page: {e}")
        sys.exit(1)
    return browser, page


def find_input_box(page):
    for sel in ["textarea >> nth=0", "textarea", "[contenteditable='true']"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=5000)
            print(f"  Input box  : {sel}")
            return loc
        except PlaywrightTimeout:
            continue
    sys.exit("ERROR: Input box not found.")


def find_button(page):
    for sel in ["button:has-text('Transliterate')", "[type='submit']", "button.btn-primary"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=5000)
            print(f"  Button     : {sel}")
            return loc
        except PlaywrightTimeout:
            continue
    sys.exit("ERROR: Transliterate button not found.")


def find_output_box(page):
    for sel in ["textarea >> nth=1", "[readonly]", "[placeholder*='inhala']"]:
        try:
            loc = page.locator(sel).first
            loc.wait_for(state="visible", timeout=3000)
            print(f"  Output box : {sel}")
            return loc
        except PlaywrightTimeout:
            continue
    print("  Output box : NOT FOUND")
    return None


def read_output(out_loc, page):
    try:
        text = out_loc.input_value(timeout=1000)
        if text and text.strip():
            return text.strip()
    except Exception:
        pass
    try:
        text = out_loc.inner_text(timeout=1000)
        if text and text.strip():
            return text.strip()
    except Exception:
        pass
    try:
        areas = page.locator("textarea").all()
        if len(areas) >= 2:
            t = areas[1].input_value(timeout=1000)
            if t and t.strip():
                return t.strip()
    except Exception:
        pass
    return ""


def clear_and_type(page, inp_loc, out_loc, text, delay_ms):
    # Clear input
    inp_loc.click()
    time.sleep(0.2)
    page.keyboard.press("Control+a")
    time.sleep(0.1)
    page.keyboard.press("Delete")
    time.sleep(0.1)
    page.keyboard.press("Control+a")
    page.keyboard.press("Backspace")
    time.sleep(0.3)

    # Try to clear output box too
    if out_loc is not None:
        try:
            out_loc.click()
            time.sleep(0.1)
            page.keyboard.press("Control+a")
            page.keyboard.press("Delete")
            time.sleep(0.1)
        except Exception:
            pass
        inp_loc.click()
        time.sleep(0.2)

    inp_loc.type(text, delay=delay_ms)
    time.sleep(0.3)


def wait_for_new_output(page, out_loc, output_before, wait_ms):
    # Wait for button to return to "Transliterate"
    try:
        page.wait_for_function(
            "() => Array.from(document.querySelectorAll('button'))"
            ".some(b => b.textContent.trim() === 'Transliterate')",
            timeout=wait_ms + 8000
        )
    except PlaywrightTimeout:
        pass

    time.sleep(1.0)

    if out_loc is None:
        return ""

    # Poll until output changes from what it was before clicking
    deadline = time.time() + 12
    while time.time() < deadline:
        current = read_output(out_loc, page)
        if current and current != output_before:
            return current
        time.sleep(0.5)

    return read_output(out_loc, page)


def main():
    args = parse_args()

    print(f"\n{'='*60}")
    print("Singlish Transliteration Test Automation")
    print(f"{'='*60}")
    print(f"Excel : {args.excel}")
    print(f"URL   : {args.url}")

    wb, ws, header_row, col_map = load_workbook_data(args.excel)
    data_rows = get_data_rows(ws, header_row, col_map)

    if not data_rows:
        sys.exit("ERROR: No data rows found.")

    print(f"Test cases found : {len(data_rows)}\n")

    passed = failed = 0

    with sync_playwright() as pw:
        browser, page = setup_page(pw, args.url, args.slow_mo_ms, args.headless)

        print("Locating page elements...")
        inp = find_input_box(page)
        btn = find_button(page)
        out = find_output_box(page)
        print("Ready! Starting tests...\n" + "="*60)

        previous_output = read_output(out, page) if out else ""

        for idx, (row_num, singlish, expected) in enumerate(data_rows, start=1):
            print(f"\n[{idx:3}/{len(data_rows)}] Row {row_num}")
            print(f"  Input    : {singlish[:70]}")

            try:
                clear_and_type(page, inp, out, singlish, args.type_delay_ms)
                output_before_click = read_output(out, page) if out else ""
                btn.click()
                actual = wait_for_new_output(page, out, output_before_click, args.wait_ms)
                previous_output = actual

                status = "PASS" if actual.strip() == expected.strip() else "FAIL"
                if status == "PASS":
                    passed += 1
                else:
                    failed += 1

                print(f"  Expected : {expected[:70]}")
                print(f"  Actual   : {actual[:70]}")
                print(f"  Status   : {status}")

                ws.cell(row=row_num, column=col_map["actual"]).value = actual
                ws.cell(row=row_num, column=col_map["status"]).value = status

            except Exception as e:
                print(f"  ERROR    : {e}")
                ws.cell(row=row_num, column=col_map["actual"]).value = f"ERROR: {e}"
                ws.cell(row=row_num, column=col_map["status"]).value = "ERROR"
                failed += 1
                previous_output = ""

            if idx % args.save_every == 0:
                wb.save(args.excel)

        wb.save(args.excel)

        print(f"\n{'='*60}")
        print(f"  PASS  : {passed}")
        print(f"  FAIL  : {failed}")
        print(f"  Total : {passed + failed}")
        print(f"  Saved : {args.excel}")
        print(f"{'='*60}\n")

        if args.keep_open:
            print("Browser kept open. Press ENTER to close...")
            input()

        browser.close()


if __name__ == "__main__":
    main()
